# -*- coding: utf-8 -*-
"""ดึงรายการจริงจากชีต Simulation พื้นที่จำลอง → seed.json ของ FinFlow"""
import json, sys, re
import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
SRC = r"C:\Users\Zbook Firefly 14 G8\Desktop\Untitled spreadsheet.xlsx"
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["Simulation พื้นที่จำลอง"]

MONTHS = [f"2026-{m:02d}" for m in range(1, 12)]      # ม.ค.–พ.ย. 2026
COL0 = 2                                              # คอลัมน์ B = ม.ค.

def num(v):
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    return None

def row_values(r):
    return [num(ws.cell(r, COL0 + i).value) for i in range(len(MONTHS))]

def label(r):
    v = ws.cell(r, 1).value
    return str(v).strip() if v else ""

# ── รายรับ (แถว 3–6) ──
INCOME_ROWS = [3, 4, 5, 6]

# ── รายจ่าย: (แถว, ประเภท, จ่ายผ่าน) ──
# หมวดตามที่ใช้จริง: fixed=ตัดไม่ได้ · variable=ปรับได้ · debt=หนี้/บัตร · saving=เงินออกไปเก็บ
EXPENSE_ROWS = [
    (8,  "variable", None,          "ส่งน้อง"),
    (9,  "fixed",    None,          "ภาษี"),
    (10, "debt",     None,          "กยศ."),
    (11, "fixed",    None,          "ค่าไฟ"),
    (12, "fixed",    None,          "ค่าน้ำ"),
    (13, "fixed",    None,          "3BB"),
    (14, "fixed",    None,          "ค่าห้อง"),
    (15, "variable", None,          "ค่ากิน (เงินสด)"),
    (16, "debt",     "cardx",       "ค่ากิน (บัตร CardX)"),
    (17, "debt",     "cardx",       "— Suunto"),
    (18, "debt",     "cardx",       "— iPhone Air"),
    (19, "debt",     "cardx",       "— ประกัน AIA"),
    (20, "debt",     "cardx",       "— ค่าอื่นๆ"),
    (21, "variable", "shopeepay",   "Shopee"),
    (22, "variable", "truewallet",  "True Wallet"),
    (23, "variable", "truewallet",  "— 7-Eleven"),
    (24, "variable", "truewallet",  "— Makro"),
    (25, "debt",     "sosmart",     "TTB So Smart"),
    (26, "fixed",    "sosmart",     "— ค่าที่จอดรถ"),
    (27, "fixed",    "sosmart",     "— ค่าประกันรถ"),
    (28, "variable", "sosmart",     "— ค่าซ่อมอื่นๆ"),
    (29, "saving",   None,          "ออมโหด"),
    (33, "debt",     "ttbflash",    "ผ่อน TTB Flash"),
]

ONETIME_ROW = 31          # เงินเม็ดพิเศษ (คาดว่าจะได้)

# วันครบกำหนดจ่าย จากตาราง "กำหนดชำระเครดิต" (แถว 40–49)
DUE_DAY = {
    "ค่าไฟ": 5, "ค่าห้อง": 14, "3BB": 25,
    "ค่ากิน (บัตร CardX)": 11, "Shopee": 25, "True Wallet": 25,
    "TTB So Smart": 7, "ผ่อน TTB Flash": 13, "ออมโหด": 5,
    "— ค่าที่จอดรถ": 1,
}

months = []
for i, key in enumerate(MONTHS):
    months.append({"id": key, "incomes": [], "expenses": [], "oneTimes": []})

for r in INCOME_ROWS:
    name = label(r) or f"รายรับแถว {r}"
    for i, v in enumerate(row_values(r)):
        if v:
            months[i]["incomes"].append({"name": name, "amount": v})

for r, typ, card, name in EXPENSE_ROWS:
    for i, v in enumerate(row_values(r)):
        if v:
            e = {"name": name, "amount": v, "type": typ}
            if card:
                e["card"] = card
            if name in DUE_DAY:
                e["dueDay"] = DUE_DAY[name]
            months[i]["expenses"].append(e)

for i, v in enumerate(row_values(ONETIME_ROW)):
    if v:
        months[i]["oneTimes"].append({"name": "เงินก้อนพิเศษ (คาดว่าจะได้)", "amount": v})

# เดือนที่ผ่านมาแล้วถือเป็นตัวเลขจริง ที่เหลือเป็นการคาดการณ์
# (ชีตทำถึง พ.ค. แล้วสูตรพัง จึงใช้ พ.ค. เป็นเส้นแบ่ง)
ACTUAL_UNTIL = "2026-05"
for m in months:
    m["status"] = "actual" if m["id"] <= ACTUAL_UNTIL else "predicted"

seed = {
    "startBalance": 500,
    "months": months,
    "cards": [
        {"id": "cardx",      "name": "SCB CardX",   "limit": 30000,  "used": 10000, "statementDay": 25, "dueDay": 11},
        {"id": "ttbflash",   "name": "TTB Flash",   "limit": 30000,  "used": 14000, "statementDay": 30, "dueDay": 13},
        {"id": "truewallet", "name": "True Wallet", "limit": 10000,  "used": 2000,  "statementDay": 15, "dueDay": 25},
        {"id": "shopeepay",  "name": "ShopeePay",   "limit": 19000,  "used": 3000,  "statementDay": 25, "dueDay": 5},
        {"id": "sosmart",    "name": "TTB So Smart","limit": 20000,  "used": 10000, "statementDay": 27, "dueDay": 7},
        {"id": "cardxflex",  "name": "CardX Flex",  "limit": 162000, "used": 0,     "statementDay": 25, "dueDay": 11},
    ],
    "buckets": [
        {"id": "dream", "name": "TTB Dream Save", "balance": 5000,  "monthlyTarget": 5000},
        {"id": "gold",  "name": "บัญชีทอง",        "balance": 20000, "monthlyTarget": 0},
        {"id": "cash",  "name": "เงินสด",          "balance": 11000, "monthlyTarget": 0},
        {"id": "make",  "name": "Make by KBank",   "balance": 10000, "monthlyTarget": 0},
    ],
    "savingsGoal": 250000,
    "people": [{"id": "me", "name": "กาย"}, {"id": "biew", "name": "พี่บิว"}],
    # ยอดหารครึ่งค้างจากชีต ประจำเดือน 2025 (ล่าสุด ธ.ค.)
    "splits": [
        {"name": "ค่าไฟ พ.ย.",     "total": 1088.35, "paidBy": "biew"},
        {"name": "ค่าน้ำ พ.ย.",    "total": 587.05,  "paidBy": "biew"},
        {"name": "ค่าเน็ต 3BB พ.ย.","total": 642.00,  "paidBy": "me"},
        {"name": "Sukishi",        "total": 976.00,  "paidBy": "me"},
    ],
}

out = r"C:\Users\Zbook Firefly 14 G8\Desktop\SummaryMoney\data\seed.json"
import os
os.makedirs(os.path.dirname(out), exist_ok=True)
with open(out, "w", encoding="utf-8") as f:
    json.dump(seed, f, ensure_ascii=False, indent=1)

tot_in = sum(len(m["incomes"]) for m in months)
tot_ex = sum(len(m["expenses"]) for m in months)
print(f"เดือน {len(months)} · รายรับ {tot_in} รายการ · รายจ่าย {tot_ex} รายการ")
print("ตัวอย่าง ก.พ.:", json.dumps(months[1], ensure_ascii=False)[:400])
